"""Simulations batch a partir de Fichier_Tir.xlsx : ajuste epaisseur
(length) et intensite (Imax) par tir, sauvegarde la vitesse de face
arriere de chaque tir, et compare le pic simule au VSL max experimental.
"""
import os
from dataclasses import replace

import numpy as np
from openpyxl import load_workbook

import control_file as cf
from lsp1d import postproc, solver
from lsp1d.loading import LaserPulseParams

SHOTS = [15, 16, 18, 19, 13, 6, 7]
DX_TARGET = 0.208  # um, meme finesse que le run de reference valide contre Radioss
T_MAX_PER_UM = 0.4  # ns/um, meme ratio que le defaut (80ns / 200um)
N_OUTPUT_SNAPSHOTS = 160  # meme ordre de grandeur que le defaut (80ns / 0.5ns)

wb = load_workbook("Fichier_Tir.xlsx", data_only=True)
ws = wb["Feuil1"]
rows = list(ws.iter_rows(min_row=2, values_only=True))

shots_data = {}
for r in rows:
    numero = r[2]
    if numero in SHOTS:
        shots_data[numero] = {
            "epaisseur_mm": r[5],
            "I_GWcm2": r[14],
            "duree_ns": r[13],
            "vsl_exp": r[15],
        }

os.makedirs("tirs", exist_ok=True)

results = []
for num in SHOTS:
    d = shots_data[num]
    length_um = d["epaisseur_mm"] * 1000.0
    Imax = d["I_GWcm2"]
    Tpul = d["duree_ns"]

    n_cells = max(200, round(length_um / DX_TARGET))
    t_max = T_MAX_PER_UM * length_um
    output_dt = t_max / N_OUTPUT_SNAPSHOTS

    cfg = replace(cf.SIM_CONFIG, length=length_um, n_cells=n_cells, t_max=t_max, output_dt=output_dt)
    pulses = [LaserPulseParams(t_start=0.0, Imax=Imax, Tpul=Tpul, T0=cf.T0)]

    mesh = solver.make_uniform_mesh(cfg.length, cfg.n_cells, cf.MATERIAL.rho0)
    state = solver.initial_state(mesh, cf.MATERIAL)
    print(f"Tir {num}: e={d['epaisseur_mm']}mm I={Imax:.2f}GW/cm2 "
          f"n_cells={n_cells} t_max={t_max:.1f}ns ...", flush=True)
    history = solver.run(mesh, state, cf.MATERIAL, cfg, pulses)

    prefix = f"tirs/tir_{num:02d}"
    postproc.save_free_surface_velocity(history, prefix + "_free_surface_velocity.csv")
    postproc.save_free_surface_velocity_excel(history, prefix + "_free_surface_velocity.xlsx")

    t_fs, u_fs = postproc.free_surface_velocity(history)
    u_fs_ms = np.abs(u_fs) * 1000.0
    vsl_sim = u_fs_ms.max()

    results.append((num, d["epaisseur_mm"], Imax, d["vsl_exp"], vsl_sim))
    print(f"  -> VSL simule = {vsl_sim:.1f} m/s (exp = {d['vsl_exp']} m/s)", flush=True)

print()
print(f"{'Tir':>4} {'e(mm)':>7} {'I(GW/cm2)':>10} {'VSL exp':>9} {'VSL sim':>9} {'ecart %':>8}")
for num, e, I, vexp, vsim in results:
    ecart = 100 * (vsim - vexp) / vexp
    print(f"{num:>4} {e:>7.2f} {I:>10.2f} {vexp:>9.0f} {vsim:>9.1f} {ecart:>7.1f}%")

import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt

nums = [r[0] for r in results]
vexps = [r[3] for r in results]
vsims = [r[4] for r in results]
x = np.arange(len(nums))
fig, ax = plt.subplots()
ax.bar(x - 0.2, vexps, width=0.4, label="VSL exp.")
ax.bar(x + 0.2, vsims, width=0.4, label="VSL Code1D")
ax.set_xticks(x)
ax.set_xticklabels([f"Tir {n}" for n in nums])
ax.set_ylabel("VSL max (m/s)")
ax.legend()
fig.savefig("tirs/comparaison_vsl.png", dpi=150)
print("-> tirs/comparaison_vsl.png")
