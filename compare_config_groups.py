"""Par configuration d'essai (diametre laser x epaisseur cible x regime
d'intensite), superpose UNE courbe Radioss de reference (DEUX si les tirs
du groupe couvrent une intensite trop dispersee) et TOUTES les courbes
experimentales disponibles du groupe : objectif = voir la dispersion des
essais autour d'une simulation de reference, pas comparer tir a tir.

Ne lance aucune simulation Code1D : uniquement Radioss (deja calcule,
disponible pour les 22 tirs, tirs/campagne_HERA_simu.h5) + experience
(disponible pour un sous-ensemble, tirs/Resultat exp/VSLXX.txt).
"""
import os

import h5py
import numpy as np
import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt

from openpyxl import load_workbook

# (diametre_mm, epaisseur_mm, regime, [tirs]) — Tableaux 4 et 5
GROUPS = [
    (1, 0.2, "I basse", [32, 35]),
    (1, 0.2, "I haute", [38, 39, 47]),
    (1, 1.0, "I basse", [33, 40]),
    (1, 1.0, "I haute", [34, 41, 46]),
    (2, 0.2, "I basse", [0]),
    (2, 0.2, "I haute", [17, 18, 19]),
    (2, 0.5, "I basse", [4, 5]),
    (2, 0.5, "I haute", [15, 16]),
    (2, 1.0, "I basse", [6, 7]),
    (2, 1.0, "I haute", [13]),
    (2, 2.0, "I haute", [20]),
]

EXP_DIR = "tirs/Resultat exp"
H5_PATH = "tirs/campagne_HERA_simu.h5"
SPREAD_THRESHOLD = 1.25  # I_max/I_min au-dela duquel on montre 2 courbes Radioss
SIGN_FLIP_EXP = {46}  # tirs ou le signe VISAR brut est inverse (verifie visuellement)

wb = load_workbook("Fichier_Tir.xlsx", data_only=True)
ws = wb["Feuil1"]
I_by_shot = {r[2]: r[14] for r in ws.iter_rows(min_row=2, values_only=True) if r[2] is not None}


def onset_time(t, v, frac=0.1):
    vmax = np.max(v)
    idx = np.argmax(v >= frac * vmax)
    return t[idx]


def align_by_xcorr(t_ref, v_ref, t_exp, v_exp, dt=0.2):
    """Decalage a appliquer a t_exp pour maximiser la correlation avec la
    courbe de reference (forme entiere du signal, robuste au bruit —
    contrairement a un seuil sur un seul point, cf. tirs 34/41/46 ou le
    bruit depasse 10% du pic avant l'arrivee du choc)."""
    t0 = min(t_ref.min(), t_exp.min())
    t1 = max(t_ref.max(), t_exp.max())
    grid = np.arange(t0, t1, dt)
    ref_g = np.interp(grid, t_ref, v_ref, left=0.0, right=0.0)
    exp_g = np.interp(grid, t_exp, v_exp, left=0.0, right=0.0)
    corr = np.correlate(ref_g, exp_g, mode="full")
    lag = np.argmax(corr) - (len(exp_g) - 1)
    return lag * dt


def reference_shots(shots):
    """1 tir (le plus proche de la moyenne) ou 2 (min/max) selon la dispersion de I."""
    if len(shots) == 1:
        return shots
    I = np.array([I_by_shot[s] for s in shots])
    if I.max() / I.min() > SPREAD_THRESHOLD:
        return [shots[int(np.argmin(I))], shots[int(np.argmax(I))]]
    return [shots[int(np.argmin(np.abs(I - I.mean())))]]


with h5py.File(H5_PATH, "r") as h5f:
    for diam, epaisseur, regime, shots in GROUPS:
        fig, ax = plt.subplots()
        colors = plt.rcParams["axes.prop_cycle"].by_key()["color"]

        ref_shots = reference_shots(shots)
        radioss_curves = {}
        for j, num in enumerate(ref_shots):
            radioss = h5f[f"Shot_{num}/DataAnalysis/velocity_simu"][:]
            t_rad = radioss[:, 0] * 1e9
            v_rad = radioss[:, 1] * 1000.0  # deja bien signee dans le HDF5 (pas d'abs : garde le rebond negatif)
            radioss_curves[num] = (t_rad, v_rad)
            ax.plot(t_rad, v_rad, color="black",
                     linestyle="-" if j == 0 else "--",
                     linewidth=2,
                     label=f"Radioss (tir {num}, I={I_by_shot[num]:.0f} GW/cm2)")

        t_ref0, v_ref0 = next(iter(radioss_curves.values()))

        n_exp = 0
        for i, num in enumerate(shots):
            exp_path = f"{EXP_DIR}/VSL{num}.txt"
            if not os.path.exists(exp_path):
                continue
            n_exp += 1
            color = colors[i % len(colors)]
            exp = np.loadtxt(exp_path, delimiter=",")
            t_exp = exp[:, 0] * 1e9
            v_exp = exp[:, 1] * 1000.0
            if num in SIGN_FLIP_EXP:
                v_exp = -v_exp
            # recale sur la courbe Radioss du meme tir si dispo (2 refs
            # possibles), sinon sur la premiere reference du groupe
            t_ref, v_ref = radioss_curves.get(num, (t_ref0, v_ref0))
            shift = align_by_xcorr(t_ref, v_ref, t_exp, v_exp)
            ax.plot(t_exp + shift, v_exp, color=color,
                     linewidth=1, alpha=0.85,
                     label=f"Exp. tir {num} (I={I_by_shot[num]:.0f} GW/cm2)")

        ax.set_xlabel("t [ns]")
        ax.set_ylabel("Vitesse face arriere [m/s]")
        ax.set_title(f"Diametre {diam}mm — epaisseur {epaisseur}mm — {regime}")
        ax.legend(fontsize=8)

        fname = f"tirs/groupe_d{diam}mm_e{epaisseur}mm_{regime.replace(' ', '')}.png"
        fig.savefig(fname, dpi=150)
        plt.close(fig)
        print(f"{diam}mm / {epaisseur}mm / {regime} : Radioss={ref_shots}, "
              f"exp disponibles={n_exp}/{len(shots)} -> {fname}")
