"""Superpose la vitesse de face arriere Code1D (run_output.npz) et la
courbe Radioss de reference (DataRADIOSS.xlsx) sur un meme graphe.

Radioss utilise une convention de signe opposee (VZ negatif quand la face
arriere s'eloigne) : on prend la valeur absolue des deux courbes pour les
comparer, comme sur ComparaisonRadioss.png.
"""
import numpy as np
from openpyxl import load_workbook

data = np.load("run_output.npz")
t_code1d = data["t"]                    # ns
u_code1d = np.abs(data["u"][:, -1]) * 1000.0  # km/s -> m/s

wb = load_workbook("DataRADIOSS.xlsx", data_only=True)
ws = wb["Feuil1"]
rows = list(ws.iter_rows(min_row=2, values_only=True))
t_radioss = np.array([float(r[0]) for r in rows]) * 1e9  # s -> ns
u_radioss = np.abs(np.array([float(r[1]) for r in rows]))

import matplotlib.pyplot as plt

fig, ax = plt.subplots()
ax.plot(t_radioss, u_radioss, label="Radioss")
ax.plot(t_code1d, u_code1d, label="Code 1D")
ax.set_xlabel("Temps (ns)")
ax.set_ylabel("Vitesse (m/s)")
ax.legend()
fig.savefig("run_output_comparison_radioss.png", dpi=150)
print("-> run_output_comparison_radioss.png")
